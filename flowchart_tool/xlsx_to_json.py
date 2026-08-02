"""
Lee una planilla de proceso (formato definido en planilla_flujograma.xlsx)
y la convierte a JSON para que el generador de PPTX la consuma.

Uso: python3 xlsx_to_json.py planilla.xlsx NombreHoja salida.json
"""
import sys
import json
import openpyxl

def main():
    xlsx_path, sheet_name, out_path = sys.argv[1], sys.argv[2], sys.argv[3]
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[sheet_name]

    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    steps = []
    for row in ws.iter_rows(min_row=2):
        values = [c.value for c in row]
        if not values[0]:
            continue
        d = dict(zip(headers, values))
        step = {
            "id": str(d.get("ID") or "").strip(),
            "nombre": str(d.get("Nombre") or "").strip(),
            "tipo": str(d.get("Tipo") or "").strip(),
            "fase": str(d.get("Fase") or "").strip(),
            "carril": str(d.get("Carril") or "").strip(),
            "siguientes": [s.strip() for s in str(d.get("Siguientes") or "").split(",") if s.strip()],
            "etiquetas_rama": [s.strip() for s in str(d.get("Etiqueta_rama") or "").split(",") if s.strip()],
            "notas": str(d.get("Notas") or "").strip(),
        }
        steps.append(step)

    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(steps, f, ensure_ascii=False, indent=2)
    print(f"OK: {len(steps)} pasos -> {out_path}")

if __name__ == "__main__":
    main()
