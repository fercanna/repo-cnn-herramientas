"""
Genera la planilla plantilla (con instrucciones) + una hoja de ejemplo cargada
con un proceso de muestra: Gestión de reclamos de clientes.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

COLUMNS = [
    ("ID", "Identificador único del paso. Usar prefijo por tipo: I=inicio/fin, P=proceso, D=decisión, Doc=documento."),
    ("Nombre", "Texto corto que describe el paso (verbo + objeto). Ej: 'Registra reclamo en sistema'."),
    ("Tipo", "Uno de: Inicio/Fin, Proceso, Decisión, Documento, Conector."),
    ("Fase", "Etapa macro del proceso (agrupa varios pasos). Ej: 'Recepción', 'Análisis'."),
    ("Carril", "Actor o área responsable del paso (define el swimlane). Ej: 'Cliente', 'Calidad'."),
    ("Siguientes", "ID(s) del/los paso(s) siguiente(s), separados por coma. Si es Decisión, poner 2 o más."),
    ("Etiqueta_rama", "Solo si Tipo=Decisión y hay más de un Siguiente: etiqueta de cada rama, en el mismo orden que Siguientes, separadas por coma. Ej: 'Sí,No'."),
    ("Notas", "Aclaración libre u opcional (referencia a anexos, sistemas, plazos, etc.)."),
]

SAMPLE_ROWS = [
    ["I1",  "Cliente presenta reclamo",        "Inicio/Fin", "Recepción", "Cliente",             "P1",     "",     ""],
    ["P1",  "Registra reclamo en sistema",     "Proceso",    "Recepción", "Atención al Cliente", "D1",     "",     "Sistema CRM"],
    ["D1",  "¿Reclamo válido?",                "Decisión",   "Análisis",  "Atención al Cliente", "P2,P6",  "Sí,No",""],
    ["P2",  "Deriva a Calidad",                "Proceso",    "Análisis",  "Atención al Cliente", "P3",     "",     ""],
    ["P3",  "Analiza causa raíz",              "Proceso",    "Análisis",  "Calidad",              "D2",     "",     ""],
    ["D2",  "¿Requiere acción correctiva?",    "Decisión",   "Resolución","Calidad",              "P4,P5",  "Sí,No",""],
    ["P4",  "Implementa acción correctiva",    "Proceso",    "Resolución","Calidad",              "P5",     "",     "Ver procedimiento AC-01"],
    ["P5",  "Informa resolución al cliente",   "Proceso",    "Resolución","Atención al Cliente",  "F1",     "",     ""],
    ["P6",  "Informa rechazo al cliente",      "Proceso",    "Análisis",  "Atención al Cliente",  "F1",     "",     ""],
    ["F1",  "Fin",                             "Inicio/Fin", "Cierre",    "Cliente",              "",       "",     ""],
]

def build():
    wb = openpyxl.Workbook()

    # --- Hoja Instrucciones ---
    ws0 = wb.active
    ws0.title = "Instrucciones"
    ws0["A1"] = "Plantilla para relevamiento de procesos → flujograma"
    ws0["A1"].font = Font(size=14, bold=True)
    ws0["A3"] = "Completá la hoja 'Proceso' con un paso por fila. Una fila por caja del flujograma."
    ws0["A4"] = "La hoja 'Ejemplo' muestra un proceso completo ya cargado (Gestión de reclamos)."
    ws0["A6"] = "Columnas:"
    ws0["A6"].font = Font(bold=True)
    row = 7
    for name, desc in COLUMNS:
        ws0.cell(row=row, column=1, value=name).font = Font(bold=True)
        ws0.cell(row=row, column=2, value=desc)
        row += 1
    ws0.column_dimensions["A"].width = 16
    ws0.column_dimensions["B"].width = 100
    for r in range(7, row):
        ws0.cell(row=r, column=2).alignment = Alignment(wrap_text=True)

    header_fill = PatternFill("solid", fgColor="1E2761")
    header_font = Font(color="FFFFFF", bold=True)

    def write_sheet(ws, rows):
        for c, (name, _desc) in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=1, column=c, value=name)
            cell.font = header_font
            cell.fill = header_fill
        for r, row_data in enumerate(rows, start=2):
            for c, val in enumerate(row_data, start=1):
                ws.cell(row=r, column=c, value=val)
        widths = [8, 32, 12, 14, 20, 14, 16, 30]
        for c, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(c)].width = w
        ws.freeze_panes = "A2"

    # --- Hoja Proceso (vacía, lista para completar) ---
    ws1 = wb.create_sheet("Proceso")
    write_sheet(ws1, [])

    # --- Hoja Ejemplo (cargada) ---
    ws2 = wb.create_sheet("Ejemplo")
    write_sheet(ws2, SAMPLE_ROWS)

    wb.save("/sessions/confident-gallant-gauss/mnt/outputs/flowchart_tool/planilla_flujograma.xlsx")
    print("OK")

if __name__ == "__main__":
    build()
